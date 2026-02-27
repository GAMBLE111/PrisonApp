from __future__ import annotations
"""
Parser for ABS Corrective Services dataset.

ABS file: "Corrective Services, Australia (Tables 1 to 17)"
- Table 1: Overview with states as columns, quarters as rows
- Table 2: Full-time custody by sex, states × time periods
"""

import openpyxl
import re
from typing import List, Dict, Any

STATE_COLUMNS = ["NSW", "Vic.", "Qld", "SA", "WA", "Tas.", "NT", "ACT", "Aust."]


def parse(filepath: str) -> List[Dict[str, Any]]:
    """Parse the Corrective Services Excel file into flat records."""
    records: List[Dict[str, Any]] = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Table 2: Full-time custody by sex and state over time
    if "Table 2" in wb.sheetnames:
        records.extend(_parse_custody_by_sex(wb["Table 2"]))

    # Table 1: Overview — persons in corrective services by state
    if "Table 1" in wb.sheetnames:
        records.extend(_parse_overview(wb["Table 1"]))

    wb.close()
    print(f"  Parsed {len(records)} corrective services records")
    return records


def _parse_custody_by_sex(sheet) -> List[Dict[str, Any]]:
    """Parse Table 2: custody numbers by sex, states as columns, time as rows."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find state column headers
    states = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        cols = []
        for j, cell in enumerate(row):
            if cell and str(cell).strip() in STATE_COLUMNS:
                cols.append((j, str(cell).strip()))
        if len(cols) >= 6:
            states = cols
            header_idx = i
            break

    if not states:
        return records

    current_sex = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Detect sex category
        label_lower = label.lower()
        if "males" in label_lower and "female" not in label_lower:
            current_sex = "Males"
            continue
        elif "females" in label_lower:
            current_sex = "Females"
            continue
        elif "persons" in label_lower:
            current_sex = "Persons"
            continue

        if not current_sex:
            continue

        # Skip percentage/rate rows
        if "%" in label or "rate" in label_lower:
            current_sex = None
            continue

        # Parse time period from label
        period = _parse_period(label)
        if not period:
            continue

        for col_idx, state_name in states:
            if state_name == "Aust.":
                continue

            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            records.append({
                "state": state_name,
                "year": period["year"],
                "quarter": period.get("quarter", "Annual"),
                "sex": current_sex,
                "legalStatus": "Full-time custody",
                "count": count,
            })

    return records


def _parse_overview(sheet) -> List[Dict[str, Any]]:
    """Parse Table 1: overview of persons in corrective services."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find state column headers
    states = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        cols = []
        for j, cell in enumerate(row):
            if cell and str(cell).strip() in STATE_COLUMNS:
                cols.append((j, str(cell).strip()))
        if len(cols) >= 6:
            states = cols
            header_idx = i
            break

    if not states:
        return records

    current_category = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        label_lower = label.lower()

        # Detect category headers
        if "males" in label_lower and "custody" in label_lower and "female" not in label_lower:
            current_category = "Males in full-time custody"
            continue
        elif "females" in label_lower and "custody" in label_lower:
            current_category = "Females in full-time custody"
            continue
        elif "persons" in label_lower and "custody" in label_lower:
            current_category = "Persons in full-time custody"
            continue
        elif "community" in label_lower:
            current_category = "Community-based corrections"
            continue

        if not current_category:
            continue

        # Parse time period
        period = _parse_period(label)
        if not period:
            continue

        sex = "Males" if "Males" in current_category else "Females" if "Females" in current_category else "Persons"
        order_type = "Full-time custody" if "custody" in current_category else "Community-based"

        for col_idx, state_name in states:
            if state_name == "Aust.":
                continue

            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            records.append({
                "state": state_name,
                "year": period["year"],
                "quarter": period.get("quarter", "Annual"),
                "sex": sex,
                "orderType": order_type,
                "count": count,
            })

    return records


def _parse_period(label: str) -> dict | None:
    """Parse a time period label like 'Sep Qtr 24', '2024', 'Jan' etc."""
    label = label.strip()

    # Full year: "2022", "2023", etc.
    if re.match(r"^\d{4}$", label):
        return {"year": label}

    # Quarter: "Sep Qtr 24", "Jun Qtr 25"
    qtr_match = re.match(r"(\w+)\s+Qtr\s+(\d{2})", label)
    if qtr_match:
        month = qtr_match.group(1)
        yr = "20" + qtr_match.group(2)
        quarter_map = {"Mar": "Q1", "Jun": "Q2", "Sep": "Q3", "Dec": "Q4"}
        qtr = quarter_map.get(month, month)
        return {"year": yr, "quarter": f"{qtr} {yr}"}

    return None


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return None
