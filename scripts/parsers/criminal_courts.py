from __future__ import annotations
"""
Parser for ABS Criminal Courts dataset.

ABS file: "Defendants finalised, Australia (Tables 1 to 6)"
- Table 1: National time series (years as columns, categories as rows)
- Table 2: State breakdown for latest year (states as columns)
- Table 3: Sex × Age × Offence cross-tab for latest year
"""

import openpyxl
from typing import List, Dict, Any

# Standard state column headers used by ABS
STATE_COLUMNS = ["NSW", "Vic.", "Qld", "SA", "WA", "Tas.", "NT", "ACT", "Aust."]


def parse(filepath: str) -> List[Dict[str, Any]]:
    """Parse the Criminal Courts Excel file into flat records."""
    records: List[Dict[str, Any]] = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Parse Table 1 for national time series
    if "Table 1" in wb.sheetnames:
        records.extend(_parse_time_series(wb["Table 1"]))

    # Parse Table 2 for state-level breakdowns (latest year)
    if "Table 2" in wb.sheetnames:
        records.extend(_parse_state_breakdown(wb["Table 2"]))

    wb.close()
    print(f"  Parsed {len(records)} criminal courts records")
    return records


def _parse_time_series(sheet) -> List[Dict[str, Any]]:
    """Parse Table 1: national time series with years as columns."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find header row with year columns
    years = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        year_cols = []
        for j, cell in enumerate(row):
            if cell and "–" in str(cell) and len(str(cell)) <= 10:
                year_cols.append((j, str(cell).strip()))
        if len(year_cols) >= 5:
            years = year_cols
            header_idx = i
            break

    if not years:
        return records

    # Parse category rows
    current_category = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Detect category headers (Sex, Age, Principal offence, etc.)
        if label in ("Sex", "Age", "Principal offence(a)", "Principal offence",
                      "Court level", "Method of finalisation", "Sentence type"):
            current_category = label.replace("(a)", "").strip()
            continue

        # Skip aggregate/summary labels
        if label.startswith("Total") or label.startswith("Mean") or label.startswith("Median"):
            continue

        # Extract counts for each year
        for col_idx, year_label in years:
            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            record = {
                "state": "Australia",
                "year": year_label,
                "count": count,
            }

            # Assign to the right field based on category
            if current_category == "Sex":
                record["sex"] = label
            elif current_category == "Age":
                record["ageGroup"] = label
            elif current_category in ("Principal offence", "Principal offence(a)"):
                record["offenceCategory"] = label
            elif current_category == "Court level":
                record["courtLevel"] = label
            elif current_category == "Method of finalisation":
                record["sentenceType"] = label
            else:
                record["sex"] = label  # fallback

            records.append(record)

    return records


def _parse_state_breakdown(sheet) -> List[Dict[str, Any]]:
    """Parse Table 2: state-level breakdown for the latest year."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find header row with state columns
    states = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        state_cols = []
        for j, cell in enumerate(row):
            if cell and str(cell).strip() in STATE_COLUMNS:
                state_cols.append((j, str(cell).strip()))
        if len(state_cols) >= 6:
            states = state_cols
            header_idx = i
            break

    if not states:
        return records

    # Determine the year from the sheet title area
    year = "2023–24"
    for row in rows[:5]:
        for cell in row:
            if cell and "–" in str(cell) and ("20" in str(cell)):
                candidate = str(cell).strip()
                # Extract year pattern like "2023-24"
                import re
                match = re.search(r"20\d{2}[–-]\d{2}", candidate)
                if match:
                    year = match.group()
                    break

    current_category = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Category headers
        if label in ("Sex", "Age", "Principal offence(a)", "Principal offence",
                      "Court level", "Method of finalisation"):
            current_category = label.replace("(a)", "").strip()
            continue

        if label.startswith("Total") or label.startswith("Mean") or label.startswith("Median"):
            continue
        if label == "All Courts":
            continue

        # Extract count for each state
        for col_idx, state_name in states:
            if state_name == "Aust.":
                continue  # skip national total, already in Table 1

            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            record = {
                "state": state_name,
                "year": year,
                "count": count,
            }

            if current_category == "Sex":
                record["sex"] = label
            elif current_category == "Age":
                record["ageGroup"] = label
            elif current_category in ("Principal offence", "Principal offence(a)"):
                record["offenceCategory"] = label
            elif current_category == "Court level":
                record["courtLevel"] = label
            elif current_category == "Method of finalisation":
                record["sentenceType"] = label
            else:
                record["sex"] = label

            records.append(record)

    return records


def _safe_int(val) -> int | None:
    """Safely convert a cell value to int."""
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return None
