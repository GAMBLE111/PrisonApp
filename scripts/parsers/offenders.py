from __future__ import annotations
"""
Parser for ABS Recorded Crime - Offenders dataset.

ABS file: "Offenders, Australia"
- Table 1: Offence categories × years (national)
- Table 5: Age × sex × years (national)
"""

import openpyxl
from typing import List, Dict, Any


def parse(filepath: str) -> List[Dict[str, Any]]:
    """Parse the Offenders Excel file into flat records."""
    records: List[Dict[str, Any]] = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Table 1: offence by year (national)
    if "Table 1" in wb.sheetnames:
        records.extend(_parse_offence_by_year(wb["Table 1"]))

    # Table 5: age by sex by year (national)
    if "Table 5" in wb.sheetnames:
        records.extend(_parse_age_sex(wb["Table 5"]))

    wb.close()
    print(f"  Parsed {len(records)} offender records")
    return records


def _parse_offence_by_year(sheet) -> List[Dict[str, Any]]:
    """Parse Table 1: offence categories by year."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find year header row
    years = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        year_cols = []
        for j, cell in enumerate(row):
            if cell and "–" in str(cell) and len(str(cell)) <= 10:
                year_cols.append((j, str(cell).strip()))
            elif cell and "-" in str(cell) and str(cell).strip().startswith("20") and len(str(cell).strip()) <= 7:
                year_cols.append((j, str(cell).strip()))
        if len(year_cols) >= 5:
            years = year_cols
            header_idx = i
            break

    if not years:
        return records

    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Only take top-level offence categories (start with 2-digit code)
        if not (len(label) > 2 and label[:2].isdigit() and label[2] == " "):
            continue

        # Skip sub-categories (3-digit codes like "011", "012")
        if len(label) > 3 and label[:3].isdigit():
            continue

        for col_idx, year_label in years:
            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            records.append({
                "state": "Australia",
                "year": year_label,
                "offenceCategory": label,
                "count": count,
            })

    return records


def _parse_age_sex(sheet) -> List[Dict[str, Any]]:
    """Parse Table 5: age by sex by year."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find year header row
    years = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        year_cols = []
        for j, cell in enumerate(row):
            s = str(cell).strip() if cell else ""
            if ("–" in s or "-" in s) and s.startswith("20") and len(s) <= 10:
                year_cols.append((j, s))
        if len(year_cols) >= 5:
            years = year_cols
            header_idx = i
            break

    if not years:
        return records

    current_sex = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""

        # Sex labels may appear in column 0 or column 1
        if not label:
            col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if col1 in ("Males", "Females"):
                current_sex = col1
            continue

        if label in ("Males", "Females"):
            current_sex = label
            continue

        if label.startswith("Total") or label.startswith("Mean") or label.startswith("Median"):
            continue

        # Age group rows contain "years"
        if "years" not in label.lower() and current_sex:
            continue

        if not current_sex:
            continue

        for col_idx, year_label in years:
            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            records.append({
                "state": "Australia",
                "year": year_label,
                "sex": current_sex,
                "ageGroup": label,
                "count": count,
            })

    return records


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return None
