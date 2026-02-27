from __future__ import annotations
"""
Parser for ABS Prisoners in Australia dataset.

ABS file: "Prisoner characteristics, Australia (Tables 1-14)"
- Table 1: Characteristics × offence categories (latest year snapshot)
- Table 2: Time series by sex, indigenous status, legal status
"""

import openpyxl
from typing import List, Dict, Any


def parse(filepath: str) -> List[Dict[str, Any]]:
    """Parse the Prisoners Excel file into flat records."""
    records: List[Dict[str, Any]] = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Table 2: time series by demographics
    if "Table 2" in wb.sheetnames:
        records.extend(_parse_time_series(wb["Table 2"]))

    # Table 1: offence × characteristics for latest year
    if "Table 1" in wb.sheetnames:
        records.extend(_parse_offence_characteristics(wb["Table 1"]))

    wb.close()
    print(f"  Parsed {len(records)} prisoner records")
    return records


def _parse_time_series(sheet) -> List[Dict[str, Any]]:
    """Parse Table 2: time series with years as rows, demographic columns."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find the column headers (row with Males, Females, etc.)
    col_headers = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        row_str = " ".join(str(c) for c in row if c).lower()
        if "males" in row_str and "females" in row_str:
            col_headers = row
            header_idx = i
            break

    if header_idx < 0:
        return records

    # Map column indices to field names
    col_map = []
    for j, cell in enumerate(col_headers):
        if not cell:
            continue
        name = str(cell).strip()
        if name == "Males":
            col_map.append((j, "sex", "Males"))
        elif name == "Females":
            col_map.append((j, "sex", "Females"))
        elif "Aboriginal" in name or "Indigenous" in name:
            col_map.append((j, "indigenousStatus", "Indigenous"))
        elif name == "Non-Indigenous":
            col_map.append((j, "indigenousStatus", "Non-Indigenous"))
        elif name == "Sentenced":
            col_map.append((j, "legalStatus", "Sentenced"))
        elif name == "Unsentenced":
            col_map.append((j, "legalStatus", "Unsentenced"))
        elif "Prior imprisonment" == name:
            col_map.append((j, "legalStatus", "Prior imprisonment"))
        elif "No prior" in name:
            col_map.append((j, "legalStatus", "No prior imprisonment"))

    # Parse year rows — only "Number" section, stop at "% change"
    in_number_section = False
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""

        if label == "Number":
            in_number_section = True
            continue
        if "% change" in label or "change" in label.lower():
            break

        if not in_number_section:
            continue

        # Year rows: "2016", "2017", etc.
        if not label.isdigit() or len(label) != 4:
            continue

        year = label

        for col_idx, field, value in col_map:
            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            records.append({
                "state": "Australia",
                "year": year,
                field: value,
                "count": count,
            })

    return records


def _parse_offence_characteristics(sheet) -> List[Dict[str, Any]]:
    """Parse Table 1: offence categories as columns, characteristics as rows."""
    records = []
    rows = list(sheet.iter_rows(values_only=True))

    # Find offence category header row
    offence_cols = []
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        cols = []
        for j, cell in enumerate(row):
            if cell and str(cell).strip().startswith("0") and " " in str(cell):
                cols.append((j, str(cell).strip()))
        if len(cols) >= 5:
            offence_cols = cols
            header_idx = i
            break

    if not offence_cols:
        return records

    # Determine year from sheet header
    year = "2025"
    for row in rows[:5]:
        for cell in row:
            if cell and "20" in str(cell):
                import re
                match = re.search(r"20\d{2}", str(cell))
                if match:
                    year = match.group()

    current_category = None
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Detect category headers
        if label in ("Sex", "Indigenous status", "Legal status", "Age",
                      "Sentence length (years)", "Prior imprisonment status"):
            current_category = label
            continue

        if label.startswith("Total") or label.startswith("Number") or label == "Imprisonment rate":
            continue

        if not current_category:
            continue

        for col_idx, offence_name in offence_cols:
            val = row[col_idx] if col_idx < len(row) else None
            count = _safe_int(val)
            if count is None or count <= 0:
                continue

            record = {
                "state": "Australia",
                "year": year,
                "offenceCategory": offence_name,
                "count": count,
            }

            if current_category == "Sex":
                record["sex"] = label
            elif current_category == "Indigenous status":
                record["indigenousStatus"] = label
            elif current_category == "Legal status":
                record["legalStatus"] = label
            elif current_category == "Age":
                record["ageGroup"] = label
            elif "Sentence" in current_category:
                record["sentenceLength"] = label

            records.append(record)

    return records


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return None
